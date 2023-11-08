import openpyxl


def find_expenses_for_numbers(input_file_path1, input_file_path2, output_file_path, numbers_file_path):
    try:
        # Завантаження файлів Excel
        input_workbook1 = openpyxl.load_workbook(input_file_path1)
        input_workbook2 = openpyxl.load_workbook(input_file_path2)
        output_workbook = openpyxl.Workbook()

        # Вибір активних аркушів (ви можете змінити назви аркушів, якщо вони відрізняються)
        input_sheet1 = input_workbook1.active
        input_sheet2 = input_workbook2.active
        output_sheet = output_workbook.active

        # Отримання номерів з файлу Excel з витратами (для першого та другого файлу)
        numbers_with_expenses1 = set(
            str(row[6]).replace(',', '') for row in input_sheet1.iter_rows(min_row=2, values_only=True))
        numbers_with_expenses2 = set(
            str(row[6]).replace(',', '') for row in input_sheet2.iter_rows(min_row=2, values_only=True))

        # Отримання даних з файлу Excel з номерами
        with open(numbers_file_path, 'r') as numbers_file:
            kyivstar_numbers = numbers_file.read().splitlines()

        # Заголовок для вихідного файлу
        output_sheet['A1'] = "Номер"
        output_sheet['B1'] = "Витрати за перший місяць"
        output_sheet['C1'] = "Витрати за другий місяць"

        # Знаходимо витрати для номерів
        row_num = 2  # Починаємо із другого рядка (після заголовка)
        for number in kyivstar_numbers:
            found_expenses1 = None
            found_expenses2 = None

            # Пошук витрат у першому файлі
            if number in numbers_with_expenses1:
                for row in input_sheet1.iter_rows(min_row=2, values_only=True):
                    if str(row[6]).replace(',', '') == number:
                        found_expenses1 = row[35]  # Витрати за перший місяць знаходяться у стовпці AJ (індекс 35)
                        break

            # Пошук витрат у другому файлі
            if number in numbers_with_expenses2:
                for row in input_sheet2.iter_rows(min_row=2, values_only=True):
                    if str(row[6]).replace(',', '') == number:
                        found_expenses2 = row[35]  # Витрати за другий місяць знаходяться у стовпці AJ (індекс 35)
                        break

            # Записуємо номер та витрати у вихідний файл
            output_sheet.cell(row=row_num, column=1, value=number)
            output_sheet.cell(row=row_num, column=2, value=found_expenses1)
            output_sheet.cell(row=row_num, column=3, value=found_expenses2)
            row_num += 1

        # Збереження результатів у вихідний файл
        output_workbook.save(output_file_path)
        print("Результати записано у вихідний файл.")

    except Exception as e:
        print(f"Помилка: {e}")

if __name__ == "__main__":
    input_file_path1 = "диф звіт травень.xlsx"  # Замініть на свій шлях до першого файлу
    input_file_path2 = "диф рахунок червень.xlsx"  # Замініть на свій шлях до другого файлу
    output_file_path = "вихідний_файл.xlsx"  # Задайте назву вихідного файлу
    numbers_file_path = "kyivstar_numbers.txt"  # Замініть на свій шлях до файлу з номерами
    find_expenses_for_numbers(input_file_path1, input_file_path2, output_file_path, numbers_file_path)